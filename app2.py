# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
from datetime import datetime

# ===================== 页面配置 =====================
st.set_page_config(page_title="问题层级处理时效与满意度分析", layout="wide")

# ===================== 全局样式 =====================
st.markdown("""
<style>
    .main { background-color: #F5F6FA; }
    h1 { color: #2B3A67; text-align: center; padding: 0.5rem 0; border-bottom: 3px solid #5B8FF9; }
    h2, h3 { color: #2B3A67; margin-top: 1.2rem; }
    div.stButton > button:first-child {
        background-color: #5B8FF9; color: white; border: none; border-radius: 8px;
        padding: 0.4rem 1.0rem;
    }
    div.stButton > button:hover { background-color: #3A6CE5; color: white; }
</style>
""", unsafe_allow_html=True)

st.title("💬 客服问题时效与满意度影响分析")

# ===================== 📚 页面导航目录 =====================
st.sidebar.title("📚 页面目录导航")
st.sidebar.markdown("### 🧩 展示模式")
show_all = st.sidebar.toggle("显示全部模块（滚动浏览）", value=True)

menu = st.sidebar.radio(
    "快速跳转到：",
    [
        "📊 整体指标概览",
        "🌟 满意度相关性分析",
        "💬 指标与满意度关系（四象限）",
        "📈 月度趋势分析",
        "🔍 问题明细散点分析",
        "🏆 问题Top榜单",
        "🌍 维度交叉热力图",
        "📤 导出分析报告",
    ],
    key="main_menu_nav"
)
st.session_state["menu"] = menu

st.markdown("""
<style>
.sidebar .sidebar-content { background-color: #F8F9FA; }
.stRadio > div { font-size: 15px; line-height: 1.6; }
</style>
""", unsafe_allow_html=True)

# ===================== 工具函数 =====================
NULL_LIKE_REGEX = {r"^[-‐-‒–—―−]+$": None, r"^(null|none|nan|NaN|NA)$": None, r"^\s*$": None}

def clean_numeric(s):
    s = s.astype(str).str.strip().replace(NULL_LIKE_REGEX, regex=True).str.replace(",", "", regex=False)
    return pd.to_numeric(s, errors="coerce")

def safe_quantile(s, q=0.9):
    s = pd.to_numeric(s, errors="coerce").dropna()
    return s.quantile(q) if len(s) > 0 else np.nan

def detect_created_col(df):
    candidates = [c for c in df.columns if "ticket_created" in c.lower() or "创建时间" in c]
    return candidates[0] if candidates else None

def ensure_time_month(df):
    created_col = detect_created_col(df)
    if created_col is None:
        st.error("❌ 未找到创建时间列（应包含 ticket_created 或 创建时间）")
        st.stop()
    df["ticket_created_datetime"] = pd.to_datetime(df[created_col], errors="coerce")
    df["month"] = df["ticket_created_datetime"].dt.to_period("M").astype(str)
    return df

def basic_clean(df):
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip().replace(NULL_LIKE_REGEX, regex=True)
    for col in ["处理时长", "评分", "message_count"]:
        if col in df.columns:
            df[col] = clean_numeric(df[col])
    return df

def group_metrics(df, level_cols, extra_dims):
    group_cols = extra_dims + level_cols
    df_valid = df.dropna(subset=["处理时长", "评分"])
    if df_valid.empty:
        return pd.DataFrame()
    grouped = (df_valid.groupby(group_cols, as_index=False)
               .agg(
                   回复次数_P90=("message_count", safe_quantile),
                   处理时长_P90=("处理时长", safe_quantile),
                   满意度_4_5占比=("评分", lambda x: (x >= 4).sum() / len(x) if len(x) > 0 else np.nan),
                   样本量=("评分", "count")
               ))
    sort_cols = [c for c in ["month", "business_line", "ticket_channel", "site_code"] if c in grouped.columns]
    return grouped.sort_values(sort_cols + level_cols)

def add_fig_to_gallery(gallery, title, fig):
    if fig is not None:
        gallery.append((title, fig))

def export_sheets_with_images(buff, sheets, filters_text, chart_gallery):
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        kaleido_ok = True
    except Exception:
        kaleido_ok = False

    with pd.ExcelWriter(buff, engine="openpyxl") as writer:
        pd.DataFrame({"筛选条件": [filters_text]}).to_excel(writer, index=False, sheet_name="筛选说明")
        for name, df in sheets.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                df.to_excel(writer, index=False, sheet_name=name)
        try:
            if kaleido_ok and len(chart_gallery) > 0:
                wb = writer.book
                ws = wb.create_sheet("图表截图")
                row_cursor = 1
                for idx, (title, fig) in enumerate(chart_gallery, start=1):
                    ws.cell(row=row_cursor, column=1, value=f"{idx}. {title}")
                    row_cursor += 1
                    try:
                        png_bytes = fig.to_image(format="png", scale=2)
                        img_stream = BytesIO(png_bytes)
                        pil_img = PILImage.open(img_stream)
                        tmp_path = f"/tmp/chart_{idx}.png"
                        pil_img.save(tmp_path)
                        xlimg = XLImage(tmp_path)
                        ws.add_image(xlimg, f"A{row_cursor}")
                        row_cursor += 30
                    except Exception:
                        ws.cell(row=row_cursor, column=1, value="（图表导出失败，可能缺少 kaleido）")
                        row_cursor += 2
        except Exception:
            pass
    buff.seek(0)

# ===================== 文件上传 =====================
uploaded = st.file_uploader("📂 上传一个或多个文件（Excel / CSV）", type=["xlsx", "csv"], accept_multiple_files=True)

if uploaded:
    dfs = []
    for f in uploaded:
        try:
            df = pd.read_excel(f) if f.name.endswith(".xlsx") else pd.read_csv(f)
            df = df.dropna(how="all").reset_index(drop=True)
            dfs.append(df)
        except Exception as e:
            st.warning(f"⚠️ 文件 {f.name} 读取失败：{e}")
    if not dfs:
        st.error("❌ 没有成功读取的文件")
        st.stop()

    df = pd.concat(dfs, ignore_index=True)
    st.success(f"✅ 已加载并合并 {len(dfs)} 个文件，共 {len(df)} 行数据。")
    st.dataframe(df.head(10), use_container_width=True)

    # ============= 数据清洗 =============
    df = ensure_time_month(df)
    df = basic_clean(df)
    for col in ["class_one", "class_two", "business_line", "ticket_channel", "site_code"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # ============= 侧边栏筛选条件 =============
    st.sidebar.header("🔎 数据筛选条件")
    min_date, max_date = df["ticket_created_datetime"].min(), df["ticket_created_datetime"].max()
    start_date, end_date = st.sidebar.date_input(
        "选择时间范围",
        value=(min_date.date() if min_date else datetime.today().date(),
               max_date.date() if max_date else datetime.today().date()),
        key="date_range_picker"
    )

    month_sel = st.sidebar.multiselect("月份", sorted(df["month"].dropna().unique()), key="month_sel")
    bl_sel = st.sidebar.multiselect("业务线", sorted(df["business_line"].dropna().unique()) if "business_line" in df.columns else [], key="bl_sel")
    ch_sel = st.sidebar.multiselect("渠道", sorted(df["ticket_channel"].dropna().unique()) if "ticket_channel" in df.columns else [], key="ch_sel")
    site_sel = st.sidebar.multiselect("国家", sorted(df["site_code"].dropna().unique()) if "site_code" in df.columns else [], key="site_sel")

    df_f = df.copy()
    if start_date and end_date:
        df_f = df_f[
            (df_f["ticket_created_datetime"] >= pd.to_datetime(start_date)) &
            (df_f["ticket_created_datetime"] <= pd.to_datetime(end_date))
        ]
    if month_sel:
        df_f = df_f[df_f["month"].isin(month_sel)]
    if bl_sel:
        df_f = df_f[df_f["business_line"].isin(bl_sel)]
    if ch_sel:
        df_f = df_f[df_f["ticket_channel"].isin(ch_sel)]
    if site_sel:
        df_f = df_f[df_f["site_code"].isin(site_sel)]

    extra_dims = [c for c in ["month", "business_line", "ticket_channel", "site_code"] if c in df_f.columns]

    # ============= 指标计算 =============
    lvl1 = group_metrics(df_f, ["class_one"], extra_dims)
    lvl2 = group_metrics(df_f, ["class_one", "class_two"], extra_dims)
    chart_gallery = []

    # ===================== 📊 整体指标概览 =====================
    if show_all or st.session_state["menu"] == "📊 整体指标概览":
        st.header("📊 整体指标概览")
        st.markdown("查看整体及各层级问题的回复次数、处理时长、满意度表现。")
        tab1, tab2 = st.tabs(["一级问题", "二级问题"])
        tab1.dataframe(lvl1, use_container_width=True)
        tab2.dataframe(lvl2, use_container_width=True)

    # ===================== 🌟 满意度相关性分析 =====================
    if show_all or st.session_state["menu"] == "🌟 满意度相关性分析":
        st.header("🌟 满意度相关性分析")
        st.markdown("计算【回复次数 / 处理时长】与【评分】的相关性，验证是否存在显著关系。")

        corr_level = st.radio("选择问题层级", ["一级问题", "二级问题"], horizontal=True, key="corr_level_radio")
        problem_field = "class_one" if corr_level == "一级问题" else "class_two"

        if problem_field not in df_f.columns:
            st.info(f"当前数据中未找到字段：{problem_field}")
        else:
            df_corr = df_f.copy().dropna(subset=[problem_field, "评分"])
            df_corr["评分"] = pd.to_numeric(df_corr["评分"], errors="coerce")
            df_corr["处理时长"] = pd.to_numeric(df_corr.get("处理时长", np.nan), errors="coerce")
            df_corr["message_count"] = pd.to_numeric(df_corr.get("message_count", np.nan), errors="coerce")

            metric_sel_corr = st.selectbox("选择用于计算相关系数的指标", ["回复次数（message_count）", "处理时长（处理时长）"], index=0, key="corr_metric_sel")
            metric_col = "message_count" if "回复次数" in metric_sel_corr else "处理时长"

            # === 新增整体相关系数 ===
            try:
                r_global = np.corrcoef(df_corr[metric_col].dropna(), df_corr["评分"].dropna())[0, 1]
                st.markdown(f"📈 **全局相关系数 r = {r_global:.3f}**（{metric_col} 与 满意度）")
            except Exception:
                r_global = np.nan

            corr_list = []
            for pb, sub in df_corr.groupby(problem_field):
                sub = sub.dropna(subset=[metric_col, "评分"])
                if len(sub) >= 5:
                    try:
                        r = np.corrcoef(sub[metric_col], sub["评分"])[0, 1]
                        corr_list.append((pb, len(sub), r))
                    except Exception:
                        pass

            if not corr_list:
                st.warning("暂无足够数据计算相关系数。")
            else:
                df_r = pd.DataFrame(corr_list, columns=["问题分类", "样本量", "相关系数"])
                df_r["相关系数"] = df_r["相关系数"].round(3)
                df_r = df_r.sort_values("相关系数", ascending=False).reset_index(drop=True)

                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("📈 正相关最高 Top5（指标升高 → 满意度升高）")
                    st.dataframe(df_r.head(5), use_container_width=True)
                with col2:
                    st.subheader("📉 负相关最高 Top5（指标升高 → 满意度下降）")
                    st.dataframe(df_r.tail(5).iloc[::-1], use_container_width=True)

                show_bar = st.checkbox("显示所有问题的相关系数条形图", value=False, key="corr_show_bar")
                if show_bar:
                    fig_bar = go.Figure()
                    fig_bar.add_trace(go.Bar(
                        x=df_r["问题分类"],
                        y=df_r["相关系数"],
                        marker_color=np.where(df_r["相关系数"] > 0, "#5B8FF9", "#E8684A"),
                        text=df_r["相关系数"],
                        textposition="outside"
                    ))
                    fig_bar.update_layout(
                        title=f"{corr_level}：{metric_sel_corr} 与 满意度 的相关系数分布",
                        xaxis_title="问题分类",
                        yaxis_title="相关系数 r",
                        xaxis_tickangle=-30,
                        plot_bgcolor="white",
                        height=600,
                        title_x=0.5,
                        title_font=dict(size=20, color="#2B3A67"),
                    )
                    st.plotly_chart(fig_bar, use_container_width=True)
                    add_fig_to_gallery(chart_gallery, f"{corr_level} 相关系数条形图 - {metric_sel_corr}", fig_bar)
                        # ===================== 💬 四象限关系 =====================
    if show_all or st.session_state["menu"] == "💬 指标与满意度关系（四象限）":
        st.header("💬 指标与满意度关系（四象限）")
        st.markdown("通过自动划分四象限，识别不同问题在【回复次数 / 处理时长】与【满意度】上的组合特征。")

        if not lvl1.empty or not lvl2.empty:
            bubble_level = st.radio("选择展示层级", ["一级问题", "二级问题"], horizontal=True, key="bubble_level_sel")
            x_metric = st.selectbox("选择横轴指标", ["处理时长_P90", "回复次数_P90"], index=1, key="bubble_x_metric_sel")
            y_metric = "满意度_4_5占比"
            problem_field = "class_one" if bubble_level == "一级问题" else "class_two"

            cur_src = lvl1 if bubble_level == "一级问题" else lvl2
            df_bubble = cur_src.copy().dropna(subset=[x_metric, y_metric])
            if df_bubble.empty or problem_field not in df_bubble.columns:
                st.warning("⚠️ 当前筛选条件下暂无可用数据。")
            else:
                df_bubble = df_bubble.groupby(problem_field, as_index=False).agg({
                    "处理时长_P90": "mean",
                    "回复次数_P90": "mean",
                    "满意度_4_5占比": "mean"
                })

                x_median = df_bubble[x_metric].median()
                y_median = df_bubble[y_metric].median()

                def quadrant_label(row):
                    if row[x_metric] >= x_median and row[y_metric] >= y_median:
                        return "高回复/高满意（积极沟通）"
                    elif row[x_metric] >= x_median and row[y_metric] < y_median:
                        return "高回复/低满意（流程瓶颈）"
                    elif row[x_metric] < x_median and row[y_metric] >= y_median:
                        return "低回复/高满意（高效解决）"
                    else:
                        return "低回复/低满意（潜在风险)"

                df_bubble["象限类型"] = df_bubble.apply(quadrant_label, axis=1)

                palette = px.colors.qualitative.Light24 + px.colors.qualitative.Set3
                unique_cats = sorted(df_bubble[problem_field].unique())
                color_map = {cat: palette[i % len(palette)] for i, cat in enumerate(unique_cats)}

                fig_bubble = go.Figure()
                for pb in sorted(df_bubble[problem_field].unique()):
                    data = df_bubble[df_bubble[problem_field] == pb]
                    fig_bubble.add_trace(go.Scatter(
                        x=data[x_metric],
                        y=data[y_metric],
                        mode="markers+text",
                        name=pb,
                        text=[pb],
                        textposition="top center",
                        marker=dict(size=14, color=color_map[pb], line=dict(width=1, color="gray"), opacity=0.9),
                        hovertemplate=f"{problem_field}: %{{text}}<br>{x_metric}: %{{x:.2f}}<br>{y_metric}: %{{y:.2f}}<extra></extra>"
                    ))

                fig_bubble.add_vline(x=x_median, line=dict(color="#666666", width=1, dash="dot"))
                fig_bubble.add_hline(y=y_median, line=dict(color="#666666", width=1, dash="dot"))

                fig_bubble.update_layout(
                    title=f"{bubble_level}：{x_metric} 与 {y_metric} 的关系（自动四象限划分）",
                    xaxis_title=x_metric, yaxis_title=y_metric,
                    plot_bgcolor="white", height=650, title_x=0.5,
                    title_font=dict(size=20, color="#2B3A67"),
                    legend=dict(orientation="h", y=1.05, x=0.5, xanchor="center")
                )
                st.plotly_chart(fig_bubble, use_container_width=True)
                add_fig_to_gallery(chart_gallery, f"{bubble_level} 四象限气泡图（{x_metric} vs {y_metric}）", fig_bubble)

                st.subheader("🔍 各象限问题明细")
                quad_choice = st.radio(
                    "选择象限类型查看对应问题：",
                    ["高回复/高满意（积极沟通）", "高回复/低满意（流程瓶颈）", "低回复/高满意（高效解决）", "低回复/低满意（潜在风险)"],
                    horizontal=True,
                    key="quadrant_choice"
                )
                df_quad = df_bubble[df_bubble["象限类型"] == quad_choice].sort_values(y_metric, ascending=False)
                if df_quad.empty:
                    st.info("该象限下暂无问题。")
                else:
                    st.dataframe(df_quad[[problem_field, "处理时长_P90", "回复次数_P90", "满意度_4_5占比"]], use_container_width=True)

    # ===================== 📈 月度趋势分析 =====================
    if show_all or st.session_state["menu"] == "📈 月度趋势分析":
        st.header("📈 指标月度趋势分析（含环比变化）")

        if "month" in df_f.columns:
            st.markdown("观察时间维度下各指标的变化趋势，并在悬停时显示环比变化。")

            trend_level = st.radio("选择层级", ["一级问题", "二级问题"], horizontal=True, key="trend_level_radio")
            trend_metric = st.selectbox("选择趋势指标", ["处理时长_P90", "回复次数_P90", "满意度_4_5占比"], index=0, key="trend_metric_sel")
            trend_dim = st.selectbox("选择分组维度", ["问题分类", "业务线", "渠道", "国家"], index=0, key="trend_dim_sel")

            problem_field = "class_one" if trend_level == "一级问题" else "class_two"
            df_trend = lvl1 if trend_level == "一级问题" else lvl2

            if df_trend.empty:
                st.info("暂无数据")
            else:
                if trend_dim == "问题分类":
                    group_field = problem_field
                elif trend_dim == "业务线":
                    group_field = "business_line"
                elif trend_dim == "渠道":
                    group_field = "ticket_channel"
                else:
                    group_field = "site_code"

                use_cols = [c for c in ["month", group_field, trend_metric] if c in df_trend.columns]
                df_trend = df_trend[use_cols].dropna(subset=[trend_metric])
                df_trend = df_trend.groupby(["month", group_field], as_index=False).mean()

                df_trend["环比变化"] = df_trend.groupby(group_field)[trend_metric].pct_change()

                sel_groups = st.multiselect(f"选择要显示的{trend_dim}", sorted(df_trend[group_field].unique()), key="trend_groups_sel")
                df_trend = df_trend[df_trend[group_field].isin(sel_groups)]

                if df_trend.empty:
                    st.warning("当前筛选条件下无数据。")
                else:
                    fig_trend = px.line(
                        df_trend,
                        x="month",
                        y=trend_metric,
                        color=group_field,
                        title=f"{trend_level}：{trend_metric} 按 {trend_dim} 的月度趋势",
                        markers=True,
                        hover_data={"环比变化": ":.2%"}
                    )
                    fig_trend.update_layout(
                        plot_bgcolor="white",
                        height=650,
                        title_x=0.5,
                        title_font=dict(size=20, color="#2B3A67"),
                        legend=dict(orientation="h", y=1.05, x=0.5, xanchor="center")
                    )
                    st.plotly_chart(fig_trend, use_container_width=True)
                    add_fig_to_gallery(chart_gallery, f"{trend_level} 月度趋势 - {trend_metric} × {trend_dim}", fig_trend)

    # ===================== 🔍 明细散点分析 =====================
    if show_all or st.session_state["menu"] == "🔍 问题明细散点分析":
        st.header("🔍 问题明细散点分析")
        st.markdown("针对单个问题，展示回复次数 / 处理时长与满意度评分之间的散点关系。")

        detail_level = st.radio("选择问题层级", ["一级问题", "二级问题"], horizontal=True, key="detail_level_radio")
        problem_field = "class_one" if detail_level == "一级问题" else "class_two"

        if problem_field not in df_f.columns:
            st.info(f"当前数据没有字段：{problem_field}")
        else:
            problem_list = sorted(df_f[problem_field].dropna().unique().tolist())
            if not problem_list:
                st.info("当前筛选下没有可选的问题分类。")
            else:
                picked_problem = st.selectbox(f"选择{detail_level}", problem_list, key="detail_pick_problem")
                x_choice = st.radio("选择横轴指标", ["回复次数（message_count）", "处理时长（处理时长）"], horizontal=True, key="detail_x_choice")
                x_col_raw = "message_count" if "回复次数" in x_choice else "处理时长"

                pts = df_f[df_f[problem_field] == picked_problem].copy().dropna(subset=[x_col_raw, "评分"])
                pts[x_col_raw] = pd.to_numeric(pts[x_col_raw], errors="coerce")
                pts["评分"] = pd.to_numeric(pts["评分"], errors="coerce")

                if pts.empty:
                    st.info("该问题分类下没有足够的数据点。")
                else:
                    r = np.corrcoef(pts[x_col_raw], pts["评分"])[0, 1]
                    st.markdown(f"📈 **相关系数 r = {r:.3f}**（{x_col_raw} 与 满意度）")

                    fig_det = px.scatter(
                        pts, x=x_col_raw, y="评分", color="ticket_channel" if "ticket_channel" in pts.columns else None,
                        trendline="ols", title=f"{picked_problem}：{x_col_raw} vs 满意度"
                    )
                    fig_det.update_layout(plot_bgcolor="white", height=600, title_x=0.5)
                    st.plotly_chart(fig_det, use_container_width=True)
                    add_fig_to_gallery(chart_gallery, f"{detail_level} 明细散点 - {picked_problem}", fig_det)

    # ===================== 🏆 问题Top榜单 =====================
    if show_all or st.session_state["menu"] == "🏆 问题Top榜单":
        st.header("🏆 问题Top榜单")
        st.markdown("识别表现最突出或最差的问题类型，用于定位改进重点。")

        level_sel = st.selectbox("选择层级", ["一级问题", "二级问题"], key="top5_level_sel")
        metric_sel = st.selectbox("选择排序指标", ["处理时长_P90", "回复次数_P90"], key="top5_metric_sel")

        cur_rank = lvl1 if level_sel == "一级问题" else lvl2
        x_col = "class_one" if level_sel == "一级问题" else "class_two"

        if cur_rank.empty:
            st.info("暂无数据。")
        else:
            df_rank = cur_rank.groupby(x_col, as_index=False).agg({
                "处理时长_P90": "mean",
                "回复次数_P90": "mean",
                "满意度_4_5占比": "mean",
                "样本量": "sum"
            })

            col1, col2 = st.columns(2)
            with col1:
                st.subheader(f"⏱️ {metric_sel.replace('_P90','')} 最高 Top5")
                st.dataframe(df_rank.sort_values(metric_sel, ascending=False).head(5), use_container_width=True)
            with col2:
                st.subheader("😞 满意度最低 Top5")
                st.dataframe(df_rank.sort_values("满意度_4_5占比", ascending=True).head(5), use_container_width=True)

    # ===================== 🌍 热力图 =====================
    if show_all or st.session_state["menu"] == "🌍 维度交叉热力图":
        st.header("🌍 维度交叉热力图")
        st.markdown("展示不同维度（业务线、渠道、国家）组合下的满意度或时效表现差异。")

        if not df_f.empty:
            x_dim = st.selectbox("选择 X 轴维度", ["business_line", "ticket_channel", "site_code"], index=0, key="hm_x_dim")
            y_dim = st.selectbox("选择 Y 轴维度", ["ticket_channel", "site_code", "business_line"], index=1, key="hm_y_dim")
            metric_sel_hm = st.radio("选择指标", ["满意度_4_5占比", "处理时长_P90", "回复次数_P90"], horizontal=True, key="hm_metric")
            if x_dim == y_dim:
                st.warning("⚠️ X 与 Y 轴不能相同。")
                df_hm = pd.DataFrame()
            else:
                df_hm = group_metrics(df_f.copy(), [], [x_dim, y_dim]).pivot(index=y_dim, columns=x_dim, values=metric_sel_hm)
                if not df_hm.empty:
                    fig_hm = go.Figure(data=go.Heatmap(
                        z=df_hm.values, x=df_hm.columns, y=df_hm.index,
                        colorscale="YlGnBu", colorbar_title=str(metric_sel_hm),
                        hovertemplate=f"{x_dim}: %{{x}}<br>{y_dim}: %{{y}}<br>{metric_sel_hm}: %{{z:.3f}}<extra></extra>"
                    ))
                    fig_hm.update_layout(
                        title=f"{metric_sel_hm} - {x_dim} × {y_dim} 热力图",
                        title_x=0.5, plot_bgcolor="white", height=700
                    )
                    st.plotly_chart(fig_hm, use_container_width=True)
                    add_fig_to_gallery(chart_gallery, f"热力图 - {metric_sel_hm}（{x_dim} × {y_dim}）", fig_hm)
        else:
            df_hm = pd.DataFrame()
    # ===================== 📋 分析结果总结 =====================
    if show_all or st.session_state["menu"] == "📋 分析结果总结":
        st.header("📋 分析结果解读与结论")
        st.markdown("根据当前数据筛选和分析结果，自动生成满意度影响分析结论。")

        # --- ① 满意度与时效总体关系 ---
        if "处理时长" in df_f.columns and "评分" in df_f.columns:
            sub = df_f.dropna(subset=["处理时长", "评分"])
            if len(sub) >= 10:
                r = np.corrcoef(sub["处理时长"], sub["评分"])[0, 1]
                if r < -0.4:
                    corr_text = f"全局相关系数 r = {r:.3f}，说明整体上呈显著负相关（时长越长满意度越低）。"
                elif r < 0.4:
                    corr_text = f"全局相关系数 r = {r:.3f}，说明整体上呈弱相关或无明显关系。"
                else:
                    corr_text = f"全局相关系数 r = {r:.3f}，说明整体上呈正相关（沟通频繁反而满意度更高）。"
            else:
                corr_text = "样本不足，无法计算相关系数。"
        else:
            corr_text = "当前数据缺少必要字段（处理时长、评分）。"

        # --- ② 四象限结构 ---
        if not lvl1.empty:
            tmp = lvl1.copy()
            x_metric, y_metric = "处理时长_P90", "满意度_4_5占比"
            x_median, y_median = tmp[x_metric].median(), tmp[y_metric].median()

            def quad(row):
                if row[x_metric] >= x_median and row[y_metric] >= y_median:
                    return "高回复/高满意（积极沟通）"
                elif row[x_metric] >= x_median and row[y_metric] < y_median:
                    return "高回复/低满意（流程瓶颈）"
                elif row[x_metric] < x_median and row[y_metric] >= y_median:
                    return "低回复/高满意（高效解决）"
                else:
                    return "低回复/低满意（潜在风险）"

            tmp["象限类型"] = tmp.apply(quad, axis=1)
            quad_counts = tmp["象限类型"].value_counts(normalize=True).mul(100).round(1).to_dict()
            quad_summary = "；".join([f"{k}：{v:.1f}%" for k, v in quad_counts.items()])
        else:
            quad_summary = "暂无可用数据"

        # --- ③ 满意度趋势 ---
        if "month" in df_f.columns and df_f["month"].nunique() >= 2:
            trend = df_f.groupby("month")["评分"].mean().sort_index()
            if len(trend) >= 2:
                diff = (trend.iloc[-1] - trend.iloc[-2]) / trend.iloc[-2]
                trend_text = f"最近两个月满意度平均变动 {diff:+.1%}。"
            else:
                trend_text = "暂无足够月度数据。"
        else:
            trend_text = "暂无时间维度数据。"

        # --- ④ 低满意问题 Top3 ---
        if not lvl1.empty:
            top_low = lvl1.sort_values("满意度_4_5占比", ascending=True).head(3)
            low_summary = "、".join([f"{r['class_one']}（{r['满意度_4_5占比']:.1%}）" for _, r in top_low.iterrows()])
        else:
            low_summary = "暂无问题分类数据。"

        # --- 汇总展示 ---
        conclusion = f"""
### 🎯 满意度影响结论摘要

1. **整体趋势：** {corr_text}  
2. **四象限结构：** {quad_summary}  
3. **满意度趋势：** {trend_text}  
4. **低满意问题：** {low_summary}

**综合判断：**
- 若处理时长显著负相关且低满意问题集中在高回复组，说明**流程效率是主要影响因素**；
- 若回复次数与满意度正相关，说明**主动沟通有助于改善体验**；
- 建议重点关注“高回复/低满意”象限问题，聚焦退款、补件、物流等慢节点。
        """

        st.markdown(conclusion)


    # ===================== 📤 导出分析报告 =====================
    if show_all or st.session_state["menu"] == "📤 导出分析报告":
        st.header("📤 导出分析报告")
        st.markdown("将当前分析结果导出为 Excel 文件（含筛选条件与图表截图）。")

        filters_text = f"时间范围: {start_date} 至 {end_date}; " \
                       f"月份: {', '.join(month_sel) if month_sel else '全部'}; " \
                       f"业务线: {', '.join(bl_sel) if bl_sel else '全部'}; " \
                       f"渠道: {', '.join(ch_sel) if ch_sel else '全部'}; " \
                       f"国家: {', '.join(site_sel) if site_sel else '全部'}"

        try:
            df_heatmap_export = df_hm.reset_index()
        except Exception:
            df_heatmap_export = pd.DataFrame()

        sheets_dict = {
            "一级问题汇总": lvl1,
            "二级问题汇总": lvl2,
            "热力图透视表": df_heatmap_export
        }

        export_buffer = BytesIO()
        export_sheets_with_images(export_buffer, sheets_dict, filters_text, chart_gallery)

        st.download_button(
            label="📥 点击下载 Excel 报告",
            data=export_buffer,
            file_name=f"满意度影响分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("请先上传数据文件（支持 CSV / XLSX，多文件可合并分析）。")

